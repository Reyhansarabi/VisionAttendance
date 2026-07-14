from PySide6.QtWidgets import (
    QWidget,
    QLabel,
    QVBoxLayout,
    QHBoxLayout,
    QApplication,
    QMessageBox,
    QDialog,
    QPushButton,
    QFileDialog,
    QComboBox,
    QAbstractItemView,
    QFrame,
    QLineEdit,
    QTimeEdit,
    QStyledItemDelegate,
)

from PySide6.QtCore import (
    Qt,
    QEvent,
    QDate,
    QTime,
    QRect,
    QMarginsF,
    QPointF,
)

from PySide6.QtGui import (
    QPainter,
    QColor,
    QPageLayout,
    QPageSize,
    QImage,
    QPolygonF,
)

from PySide6.QtPrintSupport import (
    QPrinter,
    QPrintDialog,
)

import jdatetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from theme.fonts import Fonts
from theme.colors import Colors

from gui.components.glass_card import GlassCard
from gui.components.data_table import DataTable
from gui.components.pagination import Pagination
from gui.components.table_toolbar import TableToolbar
from gui.components.local_report_chat import LocalReportChat, AnomalyCenter
from gui.components.window_toolbar import WindowToolbar

from controllers.report_controller import ReportController


# ==========================================================
# CUSTOM EXPORT COMBO BOX
# ==========================================================

class ExportComboBox(QComboBox):

    def __init__(self, parent=None):

        super().__init__(parent)

        self.setEditable(False)

        self.setMinimumHeight(46)

        self.setCursor(
            Qt.CursorShape.PointingHandCursor
        )

        self.setStyleSheet(
            """
            QComboBox {
                background-color: #F8FAFC;
                color: #334155;

                border: 1px solid #D9E3EE;
                border-radius: 10px;

                padding-left: 14px;
                padding-right: 45px;

                font-size: 11px;
                font-weight: 600;
            }

            QComboBox:hover {
                background-color: #F5F9FD;
                border: 1px solid #AFC7DC;
            }

            QComboBox:focus {
                background-color: #FFFFFF;
                border: 1px solid #5699D7;
            }

            QComboBox::drop-down {
                width: 42px;
                border: none;
                background: transparent;
            }

            QComboBox::down-arrow {
                image: none;
                width: 0px;
                height: 0px;
            }

            QComboBox QAbstractItemView {
                background-color: #FFFFFF;
                color: #334155;

                border: 1px solid #D9E3EE;
                border-radius: 10px;

                padding: 6px;

                outline: none;

                selection-background-color: #EAF3FF;
                selection-color: #1E5F94;

                font-size: 11px;
            }

            QComboBox QAbstractItemView::item {
                min-height: 38px;
                padding: 8px 12px;
                border-radius: 7px;
            }

            QComboBox QAbstractItemView::item:hover {
                background-color: #F1F7FC;
            }

            QComboBox QAbstractItemView::item:selected {
                background-color: #EAF3FF;
                color: #1E5F94;
            }
            """
        )

    # ======================================================
    # DRAW REAL TRIANGLE
    # ======================================================

    def paintEvent(self, event):

        super().paintEvent(event)

        painter = QPainter(self)

        painter.setRenderHint(
            QPainter.RenderHint.Antialiasing,
            True
        )

        # --------------------------------------------------
        # ARROW POSITION
        # --------------------------------------------------

        center_x = self.width() - 21
        center_y = self.height() // 2

        arrow_width = 9
        arrow_height = 5

        # --------------------------------------------------
        # TRIANGLE
        # --------------------------------------------------

        p1 = QPointF(
            center_x - arrow_width / 2,
            center_y - arrow_height / 2
        )

        p2 = QPointF(
            center_x + arrow_width / 2,
            center_y - arrow_height / 2
        )

        p3 = QPointF(
            center_x,
            center_y + arrow_height / 2
        )

        # --------------------------------------------------
        # COLOR
        # --------------------------------------------------

        if self.hasFocus():
            arrow_color = QColor("#5699D7")
        else:
            arrow_color = QColor("#64748B")

        painter.setPen(
            Qt.PenStyle.NoPen
        )

        painter.setBrush(
            arrow_color
        )

        painter.drawPolygon(
            QPolygonF([
                p1,
                p2,
                p3
            ])
        )

        painter.end()



# ==========================================================
# ADMIN REPORT ACTION DELEGATE
# ==========================================================

class ReportActionDelegate(QStyledItemDelegate):

    def __init__(self, table, edit_callback, delete_callback, add_callback, parent=None):
        super().__init__(parent or table)
        self.table = table
        self.edit_callback = edit_callback
        self.delete_callback = delete_callback
        self.add_callback = add_callback

    def _button_rects(self, rect, is_add=False):
        margin = 8
        gap = 6
        h = min(32, max(24, rect.height() - 14))

        if is_add:
            w = min(120, rect.width() - 2 * margin)
            y = rect.top() + (rect.height() - h) // 2
            x = rect.left() + (rect.width() - w) // 2
            return {"add": QRect(x, y, w, h)}

        w = min(58, max(42, (rect.width() - 2 * margin - gap) // 2))
        y = rect.top() + (rect.height() - h) // 2

        delete_rect = QRect(
            rect.left() + margin,
            y,
            w,
            h,
        )
        edit_rect = QRect(
            rect.right() - margin - w + 1,
            y,
            w,
            h,
        )

        return {
            "delete": delete_rect,
            "edit": edit_rect,
        }

    def paint(self, painter, option, index):
        painter.save()

        row = index.row()
        row_count = self.table.model().rowCount() if self.table.model() else 0
        is_add = row == row_count - 1

        rect = option.rect.adjusted(3, 3, -3, -3)

        painter.setPen(Qt.NoPen)
        painter.setBrush(QColor("#FCFDFE"))
        painter.drawRoundedRect(rect, 10, 10)

        rects = self._button_rects(rect, is_add=is_add)

        if is_add:
            add_rect = rects["add"]
            painter.setBrush(QColor("#EEF4FF"))
            painter.drawRoundedRect(add_rect, 8, 8)
            painter.setPen(QColor("#2F6FAB"))
            painter.drawText(
                add_rect,
                Qt.AlignCenter,
                "＋ افزودن رکورد",
            )
        else:
            edit_rect = rects["edit"]
            delete_rect = rects["delete"]

            painter.setBrush(QColor("#EEF4FF"))
            painter.setPen(Qt.NoPen)
            painter.drawRoundedRect(edit_rect, 8, 8)

            painter.setBrush(QColor("#FFF1F2"))
            painter.drawRoundedRect(delete_rect, 8, 8)

            painter.setPen(QColor("#2F6FAB"))
            painter.drawText(
                edit_rect,
                Qt.AlignCenter,
                "ویرایش",
            )

            painter.setPen(QColor("#DC4C4C"))
            painter.drawText(
                delete_rect,
                Qt.AlignCenter,
                "حذف",
            )

        painter.restore()

    def editorEvent(self, event, model, option, index):
        if event.type() != QEvent.Type.MouseButtonRelease:
            return False

        if hasattr(event, "button") and event.button() != Qt.MouseButton.LeftButton:
            return False

        row = index.row()
        row_count = model.rowCount()
        is_add = row == row_count - 1

        position = event.position().toPoint()
        rect = option.rect.adjusted(3, 3, -3, -3)
        rects = self._button_rects(rect, is_add=is_add)

        if is_add and rects["add"].contains(position):
            self.add_callback()
            return True

        if not is_add:
            if rects["edit"].contains(position):
                self.edit_callback(row)
                return True

            if rects["delete"].contains(position):
                self.delete_callback(row)
                return True

        return False


# ==========================================================
# ADMIN ATTENDANCE EDIT DIALOG
# ==========================================================

class AttendanceRecordDialog(QDialog):

    def __init__(self, users, record=None, parent=None):
        super().__init__(parent)

        self.record = record
        self.users = users

        self.setWindowTitle(
            "ویرایش رکورد حضور" if record else "افزودن رکورد حضور"
        )
        self.setModal(True)
        self.setFixedSize(430, 430)
        self.setLayoutDirection(Qt.RightToLeft)

        self.setStyleSheet(
            """
            QDialog {
                background: #FFFFFF;
            }

            QLabel {
                color: #334155;
                background: transparent;
                border: none;
                font-size: 11px;
                font-weight: 600;
            }

            QLineEdit, QTimeEdit, QComboBox {
                background: #F8FAFC;
                color: #334155;
                border: 1px solid #D9E3EE;
                border-radius: 10px;
                padding: 8px 12px;
                min-height: 38px;
                font-size: 11px;
            }

            QLineEdit:focus, QTimeEdit:focus, QComboBox:focus {
                background: #FFFFFF;
                border: 1px solid #5699D7;
            }

            QPushButton {
                border: none;
                border-radius: 9px;
                min-height: 38px;
                padding: 0 18px;
                font-size: 11px;
                font-weight: 600;
            }

            QPushButton#saveButton {
                background: #5699D7;
                color: white;
            }

            QPushButton#saveButton:hover {
                background: #4A89C2;
            }

            QPushButton#cancelButton {
                background: #F1F5F9;
                color: #475569;
            }

            QPushButton#cancelButton:hover {
                background: #E2E8F0;
            }
            """
        )

        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 24)
        layout.setSpacing(10)

        title = QLabel(self.windowTitle())
        title.setStyleSheet(
            "font-size: 16px; font-weight: bold; color: #0F172A;"
        )
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)
        layout.addSpacing(8)

        user_label = QLabel("کاربر")
        layout.addWidget(user_label)

        self.user_combo = QComboBox()
        for user in users:
            user_id = user["id"]
            name = f'{user["first_name"] or ""} {user["last_name"] or ""}'.strip()
            self.user_combo.addItem(name, user_id)
        layout.addWidget(self.user_combo)

        date_label = QLabel("تاریخ")
        layout.addWidget(date_label)

        self.date_input = QLineEdit()
        self.date_input.setPlaceholderText("مثال: 1405/05/25")
        # For a new record, default to today's date so the admin can
        # immediately save a normal attendance record without manually
        # entering the date.
        if record is None:
            self.date_input.setText(
                jdatetime.date.today().strftime("%Y/%m/%d")
            )
        layout.addWidget(self.date_input)

        time_row = QHBoxLayout()
        time_row.setSpacing(10)

        in_box = QVBoxLayout()
        in_box.setSpacing(5)
        in_box.addWidget(QLabel("ساعت ورود"))
        self.check_in = QLineEdit()
        self.check_in.setPlaceholderText("مثال: 08:15:00")
        in_box.addWidget(self.check_in)

        out_box = QVBoxLayout()
        out_box.setSpacing(5)
        out_box.addWidget(QLabel("ساعت خروج"))
        self.check_out = QLineEdit()
        self.check_out.setPlaceholderText("مثال: 16:00:00")
        out_box.addWidget(self.check_out)

        time_row.addLayout(in_box)
        time_row.addLayout(out_box)
        layout.addLayout(time_row)

        status_label = QLabel("وضعیت")
        layout.addWidget(status_label)

        self.status_combo = QComboBox()
        self.status_combo.addItems(["حاضر", "تاخیر", "غایب"])
        layout.addWidget(self.status_combo)

        layout.addStretch()

        buttons = QHBoxLayout()
        buttons.setSpacing(10)

        cancel = QPushButton("لغو")
        cancel.setObjectName("cancelButton")
        cancel.clicked.connect(self.reject)

        save = QPushButton("ذخیره تغییرات" if record else "افزودن رکورد")
        save.setObjectName("saveButton")
        save.clicked.connect(self.accept)

        buttons.addWidget(cancel)
        buttons.addWidget(save)
        layout.addLayout(buttons)

        if record is not None:
            self._load_record(record)
        elif users:
            self.user_combo.setCurrentIndex(0)

    def _load_record(self, record):
        user_id = record["user_id"]
        index = self.user_combo.findData(user_id)
        if index >= 0:
            self.user_combo.setCurrentIndex(index)

        self.date_input.setText(str(record["date"] or ""))

        if record["check_in"]:
            self.check_in.setText(str(record["check_in"]))

        if record["check_out"]:
            self.check_out.setText(str(record["check_out"]))

        status = str(record["status"] or "حاضر")
        index = self.status_combo.findText(status)
        self.status_combo.setCurrentIndex(index if index >= 0 else 0)

    def get_values(self):
        user_id = self.user_combo.currentData()
        date = self.date_input.text().strip().replace("-", "/")

        check_in = self.check_in.text().strip() or None
        check_out = self.check_out.text().strip() or None

        return {
            "user_id": user_id,
            "date": date,
            "check_in": check_in,
            "check_out": check_out,
            "status": self.status_combo.currentText(),
        }


# ==========================================================
# REPORT PAGE
# ==========================================================

class ReportPage(QWidget):

    def __init__(
        self,
        role="admin",
        user_id=None,
        parent=None
    ):

        super().__init__(parent)

        # ==================================================
        # SESSION / PERMISSIONS
        # ==================================================

        self.role = str(
            role or "user"
        ).strip().lower()

        self.user_id = user_id

        if self.role not in (
            "admin",
            "user"
        ):
            self.role = "user"

        # ==================================================
        # REPOSITORY
        # ==================================================

        self.repo = ReportController()

        # ==================================================
        # ALL DATABASE DATA
        # ==================================================

        self.all_data = []

        # ==================================================
        # FILTER STATE
        # ==================================================

        self.current_status = "all"

        self.from_date = jdatetime.date.today()
        self.to_date = jdatetime.date.today()

        # ==================================================
        # FILTERED DATA
        # ==================================================

        self.filtered_data = []
        self.all_records = []
        self.filtered_records = []
        self.page_size = 10
        self.current_page = 1
        self.page_start = 0
        self.date_filter_explicit = False

        # ==================================================
        # UI
        # ==================================================

        self.setup_ui()

        # ==================================================
        # ADMIN REPORT ACTIONS
        # ==================================================
        # The delegate must exist before load_table()/apply_filters()
        # because those methods assign it to the operations column.
        if self.role == "admin":
            self.action_delegate = ReportActionDelegate(
                self.table,
                self.edit_attendance_record,
                self.delete_attendance_record,
                self.add_attendance_record,
                self.table,
            )

        # ==================================================
        # LOAD
        # ==================================================

        self.load_table()

        # ==================================================
        # GLOBAL EVENT FILTER
        # ==================================================

        app = QApplication.instance()

        if app is not None:
            app.installEventFilter(self)

    # ==========================================================
    # UI
    # ==========================================================

    def setup_ui(self):

        self.setFocusPolicy(
            Qt.FocusPolicy.StrongFocus
        )

        layout = QVBoxLayout(self)

        layout.setContentsMargins(
            20,
            20,
            20,
            20
        )

        layout.setSpacing(12)

        # ==================================================
        # CARD
        # ==================================================

        self.card = GlassCard()

        # ==================================================
        # TITLE + SMART ANALYSIS BUTTON
        # ==================================================

        title_row = QHBoxLayout()
        title_row.setContentsMargins(0, 0, 0, 0)
        title_row.setSpacing(10)

        self.title_label = QLabel(
            "گزارش حضور و غیاب"
        )

        self.title_label.setFont(
            Fonts.heading()
        )

        self.title_label.setStyleSheet(
            f"""
            QLabel {{
                color: {Colors.TEXT};
                background: transparent;
                border: none;
            }}
            """
        )

        self.smart_analysis_button = QPushButton("✦ تحلیل هوشمند")
        self.smart_analysis_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.smart_analysis_button.setMinimumHeight(40)
        self.smart_analysis_button.setStyleSheet(
            """
            QPushButton {
                background: #EEF6FF;
                color: #1E5F94;
                border: 1px solid #BCD7EF;
                border-radius: 10px;
                padding: 0 16px;
                font-size: 12px;
                font-weight: 700;
            }
            QPushButton:hover {
                background: #E5F1FD;
                border-color: #8FBEE4;
            }
            QPushButton:pressed {
                background: #DCECF9;
            }
            """
        )
        self.smart_analysis_button.clicked.connect(self.show_smart_analysis)

        title_row.addWidget(self.title_label)
        title_row.addStretch()
        title_row.addWidget(self.smart_analysis_button)
        self.card.layout.addLayout(title_row)

        # ==================================================
        # TOOLBAR
        # ==================================================

        self.toolbar = TableToolbar()

        self.card.layout.addWidget(
            self.toolbar
        )

        # ==================================================
        # TABLE
        # ==================================================

        self.table = DataTable()

        self.card.layout.addWidget(
            self.table
        )

        # ==================================================
        # TABLE STYLE
        # ==================================================

        self.setup_report_table_style()

        # ==================================================
        # PAGINATION
        # ==================================================

        self.pagination = Pagination()

        self.card.layout.addWidget(
            self.pagination
        )

        # ==================================================
        # SMART ANALYSIS PANEL (SAME PAGE, NOT A NEW PAGE)
        # ==================================================

        self.smart_panel = GlassCard()
        self.smart_panel.setVisible(False)

        smart_header = QHBoxLayout()
        smart_header.setContentsMargins(0, 0, 0, 0)
        smart_header.setSpacing(10)

        smart_title = QLabel("تحلیل هوشمند گزارش")
        smart_title.setFont(Fonts.heading())
        smart_title.setStyleSheet(
            f"""
            QLabel {{
                color: {Colors.TEXT};
                background: transparent;
                border: none;
            }}
            """
        )

        self.smart_scope_label = QLabel("بازه انتخاب‌شده")
        self.smart_scope_label.setStyleSheet(
            "color: #64748B; background: transparent; border: none; font-size: 11px;"
        )

        self.smart_back_button = QPushButton("← بازگشت به جدول")
        self.smart_back_button.setCursor(Qt.CursorShape.PointingHandCursor)
        self.smart_back_button.setMinimumHeight(40)
        self.smart_back_button.setStyleSheet(
            """
            QPushButton {
                background: #F8FAFC;
                color: #334155;
                border: 1px solid #D9E3EE;
                border-radius: 10px;
                padding: 0 14px;
                font-size: 12px;
                font-weight: 700;
            }
            QPushButton:hover {
                background: #F1F5F9;
                border-color: #B8C9DA;
            }
            """
        )
        self.smart_back_button.clicked.connect(self.show_report_table)

        smart_header.addWidget(smart_title)
        smart_header.addWidget(self.smart_scope_label)
        smart_header.addStretch()
        smart_header.addWidget(self.smart_back_button)
        self.smart_panel.layout.addLayout(smart_header)

        self.anomaly_center = AnomalyCenter(parent=self.smart_panel)
        self.anomaly_center.setMinimumHeight(110)
        self.anomaly_center.setMaximumHeight(150)
        self.smart_panel.layout.addWidget(self.anomaly_center)

        self.local_chat = LocalReportChat(
            context_provider=self.get_ai_report_context,
            parent=self.smart_panel
        )
        self.local_chat.setMinimumHeight(230)
        self.local_chat.setMaximumHeight(310)
        self.smart_panel.layout.addWidget(self.local_chat, 1)

        layout.addWidget(self.card, 1)
        layout.addWidget(self.smart_panel, 1)

        # ==================================================
        # SIGNALS
        # ==================================================

        self.toolbar.search_changed.connect(
            self.apply_filters
        )

        self.toolbar.status_changed.connect(
            self.set_status_filter
        )

        self.toolbar.date_range_changed.connect(
            self.set_date_range
        )

        self.toolbar.refresh_clicked.connect(
            self.refresh_table
        )

        self.toolbar.print_clicked.connect(
            self.print_report
        )

        self.pagination.page_changed.connect(
            self.set_report_page
        )

    def show_smart_analysis(self):
        self.card.setVisible(False)
        self.smart_panel.setVisible(True)
        self._update_smart_scope_label()

    def show_report_table(self):
        self.smart_panel.setVisible(False)
        self.card.setVisible(True)

    def _update_smart_scope_label(self):
        start = self.from_date.strftime("%Y/%m/%d") if self.from_date else "-"
        end = self.to_date.strftime("%Y/%m/%d") if self.to_date else "-"
        self.smart_scope_label.setText(f"بازه فعال: {start} تا {end} | {len(self.filtered_records)} رکورد")

    def get_ai_report_context(self):
        """Return the live filtered report context and active date range."""
        return {
            "records": list(self.filtered_records or []),
            "start_date": self.from_date.strftime("%Y/%m/%d") if self.from_date else None,
            "end_date": self.to_date.strftime("%Y/%m/%d") if self.to_date else None,
        }

    def set_report_page(self, page):

        total_pages = max(1, (len(self.filtered_records) + self.page_size - 1) // self.page_size)
        self.current_page = max(1, min(int(page), total_pages))
        self.page_start = (self.current_page - 1) * self.page_size
        self.render_current_page()

    def render_current_page(self):

        headers = ["نام", "نام خانوادگی", "تاریخ", "ورود", "خروج", "وضعیت"]
        display_headers = headers + (["عملیات"] if self.role == "admin" else [])

        start = self.page_start
        end = start + self.page_size
        page_records = self.filtered_records[start:end]
        page_data = [
            [
                record["first_name"] or "",
                record["last_name"] or "",
                record["date"] or "",
                record["check_in"] or "-",
                record["check_out"] or "-",
                record["status"] or "-",
            ] + ([""] if self.role == "admin" else [])
            for record in page_records
        ]

        if self.role == "admin":
            page_data.append(["", "", "", "", "", "", ""])

        self.table.set_table_data(display_headers, page_data)
        self.setup_report_table_style()

        if self.role == "admin":
            self.table.setItemDelegateForColumn(6, self.action_delegate)

    # ==========================================================
    # TABLE STYLE
    # ==========================================================

    def setup_report_table_style(self):

        table_style = """
        QTableView,
        QTableWidget {
            background-color: #FFFFFF;
            alternate-background-color: #FFFFFF;

            selection-background-color: #5699D7;
            selection-color: #FFFFFF;

            gridline-color: transparent;
            border: none;
        }

        QTableView::item,
        QTableWidget::item {
            background-color: #FFFFFF;
            color: #334155;
            border: none;
        }

        QTableView::item:hover,
        QTableWidget::item:hover {
            background-color: #F1F5F9;
            color: #334155;
        }

        QTableView::item:selected,
        QTableWidget::item:selected {
            background-color: #5699D7;
            color: #FFFFFF;
        }

        QTableView::item:selected:hover,
        QTableWidget::item:selected:hover {
            background-color: #5699D7;
            color: #FFFFFF;
        }

        QTableView::item:selected:!active,
        QTableWidget::item:selected:!active {
            background-color: #5699D7;
            color: #FFFFFF;
        }

        QTableView::item:focus,
        QTableWidget::item:focus {
            outline: none;
        }
        """

        if isinstance(
            self.table,
            QAbstractItemView
        ):

            self.table.setAlternatingRowColors(
                False
            )

            self.table.setStyleSheet(
                table_style
            )

        for view in self.table.findChildren(
            QAbstractItemView
        ):

            view.setAlternatingRowColors(
                False
            )

            view.setStyleSheet(
                table_style
            )

            view.viewport().setAttribute(
                Qt.WidgetAttribute.WA_Hover,
                True
            )

    # ==========================================================
    # STATUS FILTER
    # ==========================================================

    def set_status_filter(
        self,
        status
    ):

        self.current_status = (
            status
            if status
            else "all"
        )

        # If the user has not explicitly chosen a date range, selecting a
        # status should search all of their history instead of being limited
        # to the default latest-attendance date. The same behavior is useful
        # for Admin, whose default range is today. An explicitly chosen date
        # range remains active and combines with the status filter.
        if self.current_status != "all" and not self.date_filter_explicit and self.all_records:
            parsed_dates = [
                self.parse_date(record["date"])
                for record in self.all_records
            ]
            parsed_dates = [date for date in parsed_dates if date is not None]
            if parsed_dates:
                self.from_date = min(parsed_dates)
                self.to_date = max(parsed_dates)
                self.toolbar.from_date.blockSignals(True)
                self.toolbar.to_date.blockSignals(True)
                self.toolbar.from_date.setDate(self.from_date)
                self.toolbar.to_date.setDate(self.to_date)
                self.toolbar.from_date.blockSignals(False)
                self.toolbar.to_date.blockSignals(False)

        self.apply_filters()

    # ==========================================================
    # DATE RANGE FILTER
    # ==========================================================

    def set_date_range(
        self,
        start_date,
        end_date
    ):

        if isinstance(
            start_date,
            QDate
        ):

            start_date = jdatetime.date(
                start_date.year(),
                start_date.month(),
                start_date.day()
            )

        elif not isinstance(
            start_date,
            jdatetime.date
        ):

            try:

                start_date = jdatetime.date(
                    start_date.year,
                    start_date.month,
                    start_date.day
                )

            except Exception:

                return

        if isinstance(
            end_date,
            QDate
        ):

            end_date = jdatetime.date(
                end_date.year(),
                end_date.month(),
                end_date.day()
            )

        elif not isinstance(
            end_date,
            jdatetime.date
        ):

            try:

                end_date = jdatetime.date(
                    end_date.year,
                    end_date.month,
                    end_date.day
                )

            except Exception:

                return

        # ==================================================
        # NORMALIZE RANGE
        # ==================================================

        if start_date > end_date:

            start_date, end_date = (
                end_date,
                start_date
            )

        self.from_date = start_date
        self.to_date = end_date
        self.date_filter_explicit = True

        self.apply_filters()

    # ==========================================================
    # STATUS NORMALIZER
    # ==========================================================

    def normalize_status(
        self,
        status
    ):

        if status is None:
            return ""

        value = str(
            status
        ).strip().lower()

        mapping = {

            "present": "present",
            "حاضر": "present",

            "absent": "absent",
            "غایب": "absent",

            "late": "late",
            "تاخیر": "late",
            "تاخیر دارد": "late",

        }

        return mapping.get(
            value,
            value
        )

    # ==========================================================
    # PARSE JALALI DATE
    # ==========================================================

    def parse_date(
        self,
        value
    ):

        if not value:
            return None

        if isinstance(
            value,
            jdatetime.date
        ):

            return value

        text = str(
            value
        ).strip().replace(
            "-",
            "/"
        )

        try:

            parts = text.split("/")

            if len(parts) != 3:
                return None

            year, month, day = map(
                int,
                parts
            )

            return jdatetime.date(
                year,
                month,
                day
            )

        except (
            TypeError,
            ValueError
        ):

            return None

    # ==========================================================
    # APPLY FILTERS
    # ==========================================================

    def apply_filters(self):

        headers = [
            "نام",
            "نام خانوادگی",
            "تاریخ",
            "ورود",
            "خروج",
            "وضعیت"
        ]

        if not self.all_data:
            self.filtered_data = []
            self.filtered_records = []

            display_headers = headers + (["عملیات"] if self.role == "admin" else [])
            display_data = []

            if self.role == "admin":
                display_data.append(["", "", "", "", "", "", ""])

            self.table.set_table_data(
                display_headers,
                display_data
            )

            if self.role == "admin":
                self.table.setItemDelegateForColumn(
                    6,
                    self.action_delegate
                )

            self.current_page = 1
            self.page_start = 0
            self.pagination.set_pages(1, 1)
            start_label = self.from_date.strftime("%Y/%m/%d") if self.from_date else None
            end_label = self.to_date.strftime("%Y/%m/%d") if self.to_date else None
            if hasattr(self, "local_chat"):
                self.local_chat.update_context([], start_label, end_label)
            if hasattr(self, "anomaly_center"):
                self.anomaly_center.update_context([], start_label, end_label)
            if hasattr(self, "smart_scope_label"):
                self._update_smart_scope_label()
            return

        search_text = (
            self.toolbar.search_input
            .text()
            .strip()
            .lower()
        )

        filtered = []
        filtered_records = []

        for row, record in zip(self.all_data, self.all_records):

            if search_text:
                searchable = " ".join(
                    [
                        str(row[0]),
                        str(row[1]),
                        str(row[2]),
                        str(row[3]),
                        str(row[4]),
                        str(row[5]),
                    ]
                ).lower()

                if search_text not in searchable:
                    continue

            row_date = self.parse_date(row[2])

            if row_date is not None:
                if row_date < self.from_date:
                    continue
                if row_date > self.to_date:
                    continue

            if self.current_status != "all":
                row_status = self.normalize_status(row[5])
                if row_status != self.current_status:
                    continue

            filtered.append(row)
            filtered_records.append(record)

        self.filtered_data = filtered
        self.filtered_records = filtered_records
        count = len(filtered_records)
        total_pages = max(1, (count + self.page_size - 1) // self.page_size)
        self.current_page = 1
        self.page_start = 0
        self.pagination.set_pages(1, total_pages)
        self.render_current_page()
        start_label = self.from_date.strftime("%Y/%m/%d") if self.from_date else None
        end_label = self.to_date.strftime("%Y/%m/%d") if self.to_date else None
        if hasattr(self, "local_chat"):
            self.local_chat.update_context(self.filtered_records, start_label, end_label)
        if hasattr(self, "anomaly_center"):
            self.anomaly_center.update_context(self.filtered_records, start_label, end_label)
        if hasattr(self, "smart_scope_label"):
            self._update_smart_scope_label()

    # ==========================================================
    # LOAD TABLE
    # ==========================================================

    def load_table(self, preserve_filters=False):

        today = jdatetime.date.today()

        # ==================================================
        # ADMIN
        # ==================================================

        if self.role == "admin":

            rows = self.repo.get_attendance()

        # ==================================================
        # USER
        # ==================================================

        else:

            if self.user_id is None:

                rows = []

            else:

                rows = (
                    self.repo.get_attendance_by_user(
                        self.user_id
                    )
                )

        # ==================================================
        # ALL DATA
        # ==================================================

        self.all_records = list(rows)

        self.all_data = [

            [
                row["first_name"] or "",
                row["last_name"] or "",
                row["date"] or "",
                row["check_in"] or "-",
                row["check_out"] or "-",
                row["status"] or "-",
            ]

            for row in self.all_records

        ]

        # ==================================================
        # DEFAULT DATE / TOOLBAR DATES
        # ==================================================

        if not preserve_filters:

            self.current_status = "all"
            self.date_filter_explicit = False

            default_date = today
            if self.role == "user" and self.all_records:
                latest_date = self.parse_date(self.all_records[0]["date"])
                if latest_date is not None:
                    default_date = latest_date

            self.from_date = default_date
            self.to_date = default_date

            self.toolbar.from_date.blockSignals(True)
            self.toolbar.to_date.blockSignals(True)

            self.toolbar.from_date.setDate(default_date)
            self.toolbar.to_date.setDate(default_date)

            self.toolbar.from_date.blockSignals(False)
            self.toolbar.to_date.blockSignals(False)

        # ==================================================
        # TITLE
        # ==================================================

        if self.role == "user":

            self.title_label.setText(
                "گزارش شخصی حضور و غیاب"
            )

        else:

            self.title_label.setText(
                "گزارش حضور و غیاب"
            )

        # ==================================================
        # APPLY
        # ==================================================

        self.apply_filters()

    # ==========================================================
    # REFRESH
    # ==========================================================

    def refresh_table(self):

        self.date_filter_explicit = False
        self.toolbar.reset_filters()

        self.table.clearSelection()

        self.toolbar.search_input.clearFocus()

        self.load_table()


    # ==========================================================
    # ADMIN ATTENDANCE MANAGEMENT
    # ==========================================================

    def _validate_record_values(self, values):
        date = self.parse_date(values["date"])

        if date is None:
            QMessageBox.warning(
                self,
                "تاریخ نامعتبر",
                "تاریخ را به صورت 1405/05/25 وارد کنید."
            )
            return False

        if not values["user_id"]:
            QMessageBox.warning(
                self,
                "کاربر",
                "لطفاً یک کاربر را انتخاب کنید."
            )
            return False

        for field_name, value in (
            ("ساعت ورود", values["check_in"]),
            ("ساعت خروج", values["check_out"]),
        ):
            if value:
                try:
                    parts = [int(x) for x in value.split(":")]
                    if len(parts) not in (2, 3):
                        raise ValueError
                    hour, minute = parts[0], parts[1]
                    second = parts[2] if len(parts) == 3 else 0
                    if not (0 <= hour <= 23 and 0 <= minute <= 59 and 0 <= second <= 59):
                        raise ValueError
                except (ValueError, TypeError):
                    QMessageBox.warning(
                        self,
                        "ساعت نامعتبر",
                        f"{field_name} را به صورت HH:MM یا HH:MM:SS وارد کنید."
                    )
                    return False

        return True

    def _get_users_for_dialog(self):
        return self.repo.get_users()

    def add_attendance_record(self):
        if self.role != "admin":
            return

        users = self._get_users_for_dialog()

        if not users:
            QMessageBox.information(
                self,
                "افزودن رکورد",
                "ابتدا حداقل یک کاربر در سیستم ثبت کنید."
            )
            return

        dialog = AttendanceRecordDialog(
            users,
            parent=self
        )

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        values = dialog.get_values()

        if not self._validate_record_values(values):
            return

        result = self.repo.create_attendance_record(
            values["user_id"],
            values["date"],
            values["check_in"],
            values["check_out"],
            values["status"],
        )

        if not result.get("success"):
            if result.get("reason") == "duplicate":
                QMessageBox.warning(
                    self,
                    "رکورد تکراری",
                    "برای این کاربر در این تاریخ، رکورد حضور قبلاً ثبت شده است."
                )
            else:
                QMessageBox.warning(
                    self,
                    "خطا",
                    "ثبت رکورد انجام نشد."
                )
            return

        # Make the newly-created record immediately visible in the report,
        # even if the report previously had a different date/status/search filter.
        self.toolbar.search_input.clear()
        self.current_status = "all"
        new_date = self.parse_date(values["date"])
        if new_date is not None:
            self.from_date = new_date
            self.to_date = new_date
            self.toolbar.from_date.blockSignals(True)
            self.toolbar.to_date.blockSignals(True)
            self.toolbar.from_date.setDate(new_date)
            self.toolbar.to_date.setDate(new_date)
            self.toolbar.from_date.blockSignals(False)
            self.toolbar.to_date.blockSignals(False)

        self.load_table(preserve_filters=True)

    def edit_attendance_record(self, row_index):
        if self.role != "admin":
            return

        actual_index = self.page_start + row_index
        if actual_index < 0 or actual_index >= len(self.filtered_records):
            return

        record = self.filtered_records[actual_index]
        users = self._get_users_for_dialog()

        dialog = AttendanceRecordDialog(
            users,
            record=record,
            parent=self
        )

        if dialog.exec() != QDialog.DialogCode.Accepted:
            return

        values = dialog.get_values()

        if not self._validate_record_values(values):
            return

        success = self.repo.update_attendance_record(
            record["id"],
            values["user_id"],
            values["date"],
            values["check_in"],
            values["check_out"],
            values["status"],
        )

        if not success:
            QMessageBox.warning(
                self,
                "ویرایش انجام نشد",
                "رکورد پیدا نشد یا برای این کاربر و تاریخ رکورد تکراری وجود دارد."
            )
            return

        self.load_table(preserve_filters=True)

    def delete_attendance_record(self, row_index):
        if self.role != "admin":
            return

        actual_index = self.page_start + row_index
        if actual_index < 0 or actual_index >= len(self.filtered_records):
            return

        record = self.filtered_records[actual_index]

        name = (
            f'{record["first_name"] or ""} '
            f'{record["last_name"] or ""}'
        ).strip()

        answer = QMessageBox.question(
            self,
            "حذف رکورد",
            f"آیا از حذف رکورد حضور «{name}» در تاریخ {record['date']} مطمئن هستید؟",
            QMessageBox.StandardButton.Yes
            | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No,
        )

        if answer != QMessageBox.StandardButton.Yes:
            return

        success = self.repo.delete_attendance_record(
            record["id"]
        )

        if not success:
            QMessageBox.warning(
                self,
                "خطا",
                "حذف رکورد انجام نشد."
            )
            return

        self.load_table(preserve_filters=True)

    # ==========================================================
    # PRINT / EXPORT DIALOG
    # ==========================================================

    def print_report(self):

        if not self.filtered_data:

            QMessageBox.information(
                self,
                "خروجی گزارش",
                "داده‌ای برای خروجی وجود ندارد."
            )

            return

        # ==================================================
        # DIALOG
        # ==================================================

        dialog = QDialog(
            self
        )

        # --------------------------------------------------
        # Remove the default Windows title bar.
        # The dialog uses the project's custom WindowToolbar.
        # --------------------------------------------------
        dialog.setWindowFlags(
            Qt.WindowType.Dialog
            | Qt.WindowType.FramelessWindowHint
        )

        dialog.setWindowTitle(
            "خروجی گزارش حضور و غیاب"
        )

        dialog.setModal(
            True
        )

        dialog.setFixedSize(
            460,
            374
        )

        dialog.setStyleSheet(
            """
            QDialog {
                background: #FFFFFF;
            }

            QLabel#Title {
                color: #0F172A;
                font-size: 18px;
                font-weight: 700;

                background: transparent;
                border: none;
            }

            QLabel#Subtitle {
                color: #64748B;
                font-size: 10px;

                background: transparent;
                border: none;
            }

            QLabel#OutputLabel {
                color: #334155;
                font-size: 11px;
                font-weight: 600;

                background: transparent;
                border: none;
            }

            QLabel#InfoLabel {
                color: #64748B;
                font-size: 10px;

                background: transparent;
                border: none;
            }

            QFrame#InfoFrame {
                background: #F8FAFC;

                border: 1px solid #E2E8F0;
                border-radius: 10px;
            }

            QPushButton {
                min-height: 44px;

                border-radius: 10px;
                border: none;

                padding-left: 20px;
                padding-right: 20px;

                font-size: 11px;
                font-weight: 600;
            }

            QPushButton#ContinueButton {
                background: #5699D7;
                color: #FFFFFF;
            }

            QPushButton#ContinueButton:hover {
                background: #478CCB;
            }

            QPushButton#ContinueButton:pressed {
                background: #3F7FBA;
            }

            QPushButton#CancelButton {
                background: #F1F5F9;
                color: #475569;
            }

            QPushButton#CancelButton:hover {
                background: #E2E8F0;
            }

            QPushButton#CancelButton:pressed {
                background: #E2E8F0;
            }
            """
        )

        # ==================================================
        # CUSTOM WINDOW TOOLBAR
        # ==================================================

        root_layout = QVBoxLayout(
            dialog
        )

        root_layout.setContentsMargins(
            0,
            0,
            0,
            0
        )

        root_layout.setSpacing(
            0
        )

        window_toolbar = WindowToolbar(
            dialog,
            dialog
        )

        window_toolbar.minimize_button.hide()
        window_toolbar.maximize_button.hide()

        window_toolbar.title.setText(
            "خروجی گزارش حضور و غیاب"
        )

        root_layout.addWidget(
            window_toolbar
        )

        # ==================================================
        # MAIN CONTENT
        # ==================================================

        content_widget = QWidget()
        content_widget.setObjectName(
            "ExportContent"
        )

        content_layout = QVBoxLayout(
            content_widget
        )

        content_layout.setContentsMargins(
            28,
            24,
            28,
            24
        )

        content_layout.setSpacing(
            10
        )

        root_layout.addWidget(
            content_widget,
            1
        )

        main_layout = content_layout

        # ==================================================
        # TITLE
        # ==================================================

        title = QLabel(
            "خروجی گزارش حضور و غیاب"
        )

        title.setObjectName(
            "Title"
        )

        title.setAlignment(
            Qt.AlignmentFlag.AlignCenter
        )

        main_layout.addWidget(
            title
        )

        # ==================================================
        # SUBTITLE
        # ==================================================

        subtitle = QLabel(
            "فرمت موردنظر برای ذخیره یا چاپ گزارش را انتخاب کنید."
        )

        subtitle.setObjectName(
            "Subtitle"
        )

        subtitle.setAlignment(
            Qt.AlignmentFlag.AlignCenter
        )

        main_layout.addWidget(
            subtitle
        )

        main_layout.addSpacing(
            12
        )

        # ==================================================
        # OUTPUT LABEL
        # ==================================================

        output_label = QLabel(
            "نوع خروجی"
        )

        output_label.setObjectName(
            "OutputLabel"
        )

        main_layout.addWidget(
            output_label
        )

        # ==================================================
        # OUTPUT COMBO
        # ==================================================

        output_combo = ExportComboBox()

        output_combo.addItem(
            "چاپ با پرینتر",
            "printer"
        )

        output_combo.addItem(
            "ذخیره به صورت PDF",
            "pdf"
        )

        output_combo.addItem(
            "ذخیره به صورت JPG",
            "jpg"
        )

        output_combo.addItem(
            "خروجی Excel",
            "excel"
        )

        output_combo.setCurrentIndex(
            0
        )

        main_layout.addWidget(
            output_combo
        )

        # ==================================================
        # INFO FRAME
        # ==================================================

        info_frame = QFrame()

        info_frame.setObjectName(
            "InfoFrame"
        )

        info_layout = QHBoxLayout(
            info_frame
        )

        info_layout.setContentsMargins(
            12,
            8,
            12,
            8
        )

        info_label = QLabel(
            f"تعداد رکوردهای قابل خروجی: "
            f"{len(self.filtered_data)}"
        )

        info_label.setObjectName(
            "InfoLabel"
        )

        info_label.setAlignment(
            Qt.AlignmentFlag.AlignCenter
        )

        info_layout.addWidget(
            info_label
        )

        main_layout.addWidget(
            info_frame
        )

        main_layout.addSpacing(
            8
        )

        # ==================================================
        # BUTTONS
        # ==================================================

        buttons_layout = QHBoxLayout()

        buttons_layout.setSpacing(
            10
        )

        cancel_button = QPushButton(
            "انصراف"
        )

        cancel_button.setObjectName(
            "CancelButton"
        )

        continue_button = QPushButton(
            "ادامه"
        )

        continue_button.setObjectName(
            "ContinueButton"
        )

        cancel_button.setCursor(
            Qt.CursorShape.PointingHandCursor
        )

        continue_button.setCursor(
            Qt.CursorShape.PointingHandCursor
        )

        buttons_layout.addWidget(
            cancel_button
        )

        buttons_layout.addWidget(
            continue_button
        )

        main_layout.addLayout(
            buttons_layout
        )

        # ==================================================
        # EXPORT
        # ==================================================

        def export_report():

            output_type = (
                output_combo.currentData()
            )

            dialog.accept()

            if output_type == "printer":

                self.print_to_printer()

            elif output_type == "pdf":

                self.save_report_pdf()

            elif output_type == "jpg":

                self.save_report_jpg()

            elif output_type == "excel":

                self.save_report_excel()

        continue_button.clicked.connect(
            export_report
        )

        cancel_button.clicked.connect(
            dialog.reject
        )

        # ==================================================
        # SHOW
        # ==================================================

        dialog.exec()

    # ==========================================================
    # CREATE PRINTER
    # ==========================================================

    def create_printer(
        self,
        output_file=None
    ):

        printer = QPrinter(
            QPrinter.PrinterMode.HighResolution
        )

        page_size = QPageSize(
            QPageSize.PageSizeId.A4
        )

        page_layout = QPageLayout(
            page_size,
            QPageLayout.Orientation.Landscape,
            QMarginsF(
                12.0,
                12.0,
                12.0,
                12.0
            ),
            QPageLayout.Unit.Millimeter
        )

        printer.setPageLayout(
            page_layout
        )

        if output_file:

            printer.setOutputFormat(
                QPrinter.OutputFormat.PdfFormat
            )

            printer.setOutputFileName(
                output_file
            )

        return printer

    # ==========================================================
    # SAVE PDF
    # ==========================================================

    def save_report_pdf(self):

        default_name = (
            "گزارش_حضور_و_غیاب.pdf"
        )

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "ذخیره گزارش PDF",
            default_name,
            "PDF Files (*.pdf)"
        )

        if not file_path:
            return

        if not file_path.lower().endswith(
            ".pdf"
        ):

            file_path += ".pdf"

        printer = self.create_printer(
            file_path
        )

        painter = QPainter()

        if not painter.begin(
            printer
        ):

            QMessageBox.warning(
                self,
                "خطا",
                "امکان ایجاد فایل PDF وجود ندارد."
            )

            return

        try:

            self.draw_print_report(
                painter,
                printer
            )

        except Exception as error:

            QMessageBox.critical(
                self,
                "خطا در ساخت PDF",
                f"خطایی هنگام ساخت گزارش رخ داد:\n\n{error}"
            )

            return

        finally:

            painter.end()

        QMessageBox.information(
            self,
            "گزارش",
            "گزارش PDF با موفقیت ذخیره شد."
        )

    # ==========================================================
    # PRINT
    # ==========================================================

    def print_to_printer(self):

        printer = self.create_printer()

        dialog = QPrintDialog(
            printer,
            self
        )

        dialog.setWindowTitle(
            "چاپ گزارش حضور و غیاب"
        )

        result = dialog.exec()

        if result != QDialog.DialogCode.Accepted:
            return

        painter = QPainter()

        if not painter.begin(
            printer
        ):

            QMessageBox.warning(
                self,
                "خطا",
                "امکان شروع عملیات چاپ وجود ندارد."
            )

            return

        try:

            self.draw_print_report(
                painter,
                printer
            )

        except Exception as error:

            QMessageBox.critical(
                self,
                "خطا در چاپ",
                f"خطایی هنگام چاپ گزارش رخ داد:\n\n{error}"
            )

        finally:

            painter.end()

    # ==========================================================
    # SAVE JPG
    # ==========================================================

    def save_report_jpg(self):

        default_name = (
            "گزارش_حضور_و_غیاب.jpg"
        )

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "ذخیره گزارش JPG",
            default_name,
            "JPG Files (*.jpg *.jpeg)"
        )

        if not file_path:
            return

        if not file_path.lower().endswith(
            (".jpg", ".jpeg")
        ):

            file_path += ".jpg"

        # ==================================================
        # IMAGE SETTINGS
        # ==================================================

        dpi = 300

        width_mm = 297
        height_mm = 210

        image_width = int(
            width_mm * dpi / 25.4
        )

        image_height = int(
            height_mm * dpi / 25.4
        )

        image = QImage(
            image_width,
            image_height,
            QImage.Format.Format_RGB32
        )

        image.fill(
            QColor("#FFFFFF")
        )

        image.setDotsPerMeterX(
            int(dpi / 0.0254)
        )

        image.setDotsPerMeterY(
            int(dpi / 0.0254)
        )

        painter = QPainter()

        if not painter.begin(
            image
        ):

            QMessageBox.warning(
                self,
                "خطا",
                "امکان ایجاد تصویر گزارش وجود ندارد."
            )

            return

        try:

            self.draw_report_on_image(
                painter,
                image
            )

        except Exception as error:

            painter.end()

            QMessageBox.critical(
                self,
                "خطا در ساخت JPG",
                f"خطایی هنگام ساخت تصویر رخ داد:\n\n{error}"
            )

            return

        painter.end()

        if not image.save(
            file_path,
            "JPG",
            95
        ):

            QMessageBox.warning(
                self,
                "خطا",
                "امکان ذخیره فایل JPG وجود ندارد."
            )

            return

        QMessageBox.information(
            self,
            "گزارش",
            "گزارش JPG با موفقیت ذخیره شد."
        )

    # ==========================================================
    # SAVE EXCEL
    # ==========================================================

    def save_report_excel(self):

        if not self.filtered_data:

            QMessageBox.information(
                self,
                "خروجی Excel",
                "داده‌ای برای خروجی وجود ندارد."
            )

            return

        default_name = (
            "گزارش_حضور_و_غیاب.xlsx"
        )

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "ذخیره گزارش Excel",
            default_name,
            "Excel Files (*.xlsx)"
        )

        if not file_path:
            return

        if not file_path.lower().endswith(
            ".xlsx"
        ):

            file_path += ".xlsx"

        try:

            # ==================================================
            # WORKBOOK
            # ==================================================

            workbook = Workbook()

            worksheet = workbook.active

            worksheet.title = (
                "گزارش حضور و غیاب"
            )

            # ==================================================
            # HEADERS
            # ==================================================

            headers = [
                "نام",
                "نام خانوادگی",
                "تاریخ",
                "ورود",
                "خروج",
                "وضعیت"
            ]

            for column, header in enumerate(
                headers,
                start=1
            ):

                cell = worksheet.cell(
                    row=1,
                    column=column,
                    value=header
                )

                cell.font = Font(
                    bold=True,
                    size=11
                )

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="EAF3FF"
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # ==================================================
            # DATA
            # ==================================================

            for row_index, row in enumerate(
                self.filtered_data,
                start=2
            ):

                for column_index, value in enumerate(
                    row,
                    start=1
                ):

                    cell = worksheet.cell(
                        row=row_index,
                        column=column_index,
                        value=str(value)
                    )

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

            # ==================================================
            # BORDER
            # ==================================================

            thin_border = Border(
                left=Side(
                    style="thin",
                    color="D9E3EE"
                ),
                right=Side(
                    style="thin",
                    color="D9E3EE"
                ),
                top=Side(
                    style="thin",
                    color="D9E3EE"
                ),
                bottom=Side(
                    style="thin",
                    color="D9E3EE"
                )
            )

            for row in worksheet.iter_rows():

                for cell in row:

                    cell.border = thin_border

            # ==================================================
            # COLUMN WIDTH
            # ==================================================

            widths = {
                1: 18,
                2: 20,
                3: 15,
                4: 14,
                5: 14,
                6: 16,
            }

            for column, width in widths.items():

                worksheet.column_dimensions[
                    get_column_letter(column)
                ].width = width

            # ==================================================
            # ROW HEIGHT
            # ==================================================

            worksheet.row_dimensions[
                1
            ].height = 28

            for row_index in range(
                2,
                len(self.filtered_data) + 2
            ):

                worksheet.row_dimensions[
                    row_index
                ].height = 24

            # ==================================================
            # FREEZE HEADER
            # ==================================================

            worksheet.freeze_panes = "A2"

            # ==================================================
            # AUTO FILTER
            # ==================================================

            worksheet.auto_filter.ref = (
                f"A1:F{len(self.filtered_data) + 1}"
            )

            # ==================================================
            # SAVE
            # ==================================================

            workbook.save(
                file_path
            )

        except Exception as error:

            QMessageBox.critical(
                self,
                "خطا در خروجی Excel",
                f"ساخت فایل Excel انجام نشد:\n\n{error}"
            )

            return

        QMessageBox.information(
            self,
            "خروجی گزارش",
            "گزارش Excel با موفقیت ذخیره شد."
        )

    # ==========================================================
    # DRAW REPORT ON IMAGE
    # ==========================================================

    def draw_report_on_image(
        self,
        painter,
        image
    ):

        width = image.width()
        height = image.height()

        dpi = 300

        def mm(value):

            return int(
                value * dpi / 25.4
            )

        # ==================================================
        # MARGINS
        # ==================================================

        margin_left = mm(12)
        margin_right = mm(12)
        margin_top = mm(12)
        margin_bottom = mm(12)

        content_left = margin_left
        content_top = margin_top

        content_width = (
            width
            - margin_left
            - margin_right
        )

        content_bottom = (
            height
            - margin_bottom
        )

        # ==================================================
        # BACKGROUND
        # ==================================================

        painter.fillRect(
            QRect(
                content_left,
                content_top,
                content_width,
                content_bottom - content_top
            ),
            QColor("#FFFFFF")
        )

        # ==================================================
        # TITLE
        # ==================================================

        title_font = painter.font()

        title_font.setPointSize(17)
        title_font.setBold(True)

        painter.setFont(
            title_font
        )

        painter.setPen(
            QColor("#0F172A")
        )

        title_height = mm(12)

        title_rect = QRect(
            content_left,
            content_top,
            content_width,
            title_height
        )

        painter.drawText(
            title_rect,
            Qt.AlignmentFlag.AlignCenter,
            "گزارش حضور و غیاب"
        )

        # ==================================================
        # INFO
        # ==================================================

        info_font = painter.font()

        info_font.setPointSize(9)
        info_font.setBold(False)

        painter.setFont(
            info_font
        )

        from_text = self.from_date.strftime(
            "%Y/%m/%d"
        )

        to_text = self.to_date.strftime(
            "%Y/%m/%d"
        )

        status_text = {

            "all": "همه وضعیت‌ها",
            "present": "حاضر",
            "late": "تاخیر",
            "absent": "غایب",

        }.get(
            self.current_status,
            self.current_status
        )

        info = (
            f"بازه تاریخ: {from_text} تا {to_text}"
            f"     |     وضعیت: {status_text}"
            f"     |     تعداد: {len(self.filtered_data)}"
        )

        info_height = mm(9)

        info_rect = QRect(
            content_left,
            content_top + title_height,
            content_width,
            info_height
        )

        painter.setPen(
            QColor("#64748B")
        )

        painter.drawText(
            info_rect,
            Qt.AlignmentFlag.AlignCenter,
            info
        )

        # ==================================================
        # TABLE
        # ==================================================

        self._draw_report_table(
            painter,
            content_left,
            content_top
            + title_height
            + info_height
            + mm(8),
            content_width,
            content_bottom,
            mm,
            is_printer=False
        )

    # ==========================================================
    # DRAW PRINT REPORT
    # ==========================================================

    def draw_print_report(
        self,
        painter,
        printer
    ):

        page_rect = printer.pageRect(
            QPrinter.Unit.DevicePixel
        )

        left = page_rect.left()
        top = page_rect.top()

        width = page_rect.width()
        height = page_rect.height()

        dpi = printer.resolution()

        if dpi <= 0:
            dpi = 300

        def mm(value):

            return int(
                value * dpi / 25.4
            )

        # ==================================================
        # MARGINS
        # ==================================================

        margin_left = mm(12)
        margin_right = mm(12)
        margin_top = mm(12)
        margin_bottom = mm(12)

        content_left = (
            left + margin_left
        )

        content_top = (
            top + margin_top
        )

        content_width = (
            width
            - margin_left
            - margin_right
        )

        content_bottom = (
            top
            + height
            - margin_bottom
        )

        # ==================================================
        # BACKGROUND
        # ==================================================

        painter.fillRect(
            QRect(
                content_left,
                content_top,
                content_width,
                content_bottom - content_top
            ),
            QColor("#FFFFFF")
        )

        # ==================================================
        # TITLE
        # ==================================================

        title_font = painter.font()

        title_font.setPointSize(17)
        title_font.setBold(True)

        painter.setFont(
            title_font
        )

        painter.setPen(
            QColor("#0F172A")
        )

        title_height = mm(12)

        title_rect = QRect(
            content_left,
            content_top,
            content_width,
            title_height
        )

        painter.drawText(
            title_rect,
            Qt.AlignmentFlag.AlignCenter,
            "گزارش حضور و غیاب"
        )

        # ==================================================
        # INFO
        # ==================================================

        info_font = painter.font()

        info_font.setPointSize(9)
        info_font.setBold(False)

        painter.setFont(
            info_font
        )

        from_text = self.from_date.strftime(
            "%Y/%m/%d"
        )

        to_text = self.to_date.strftime(
            "%Y/%m/%d"
        )

        status_text = {

            "all": "همه وضعیت‌ها",
            "present": "حاضر",
            "late": "تاخیر",
            "absent": "غایب",

        }.get(
            self.current_status,
            self.current_status
        )

        info = (
            f"بازه تاریخ: {from_text} تا {to_text}"
            f"     |     وضعیت: {status_text}"
            f"     |     تعداد: {len(self.filtered_data)}"
        )

        info_height = mm(9)

        info_rect = QRect(
            content_left,
            content_top + title_height,
            content_width,
            info_height
        )

        painter.setPen(
            QColor("#64748B")
        )

        painter.drawText(
            info_rect,
            Qt.AlignmentFlag.AlignCenter,
            info
        )

        # ==================================================
        # TABLE
        # ==================================================

        self._draw_report_table(
            painter,
            content_left,
            content_top
            + title_height
            + info_height
            + mm(8),
            content_width,
            content_bottom,
            mm,
            is_printer=True,
            printer=printer
        )

    # ==========================================================
    # DRAW TABLE
    # ==========================================================

    def _draw_report_table(
        self,
        painter,
        content_left,
        table_top,
        content_width,
        content_bottom,
        mm,
        is_printer=False,
        printer=None
    ):

        headers = [
            "نام",
            "نام خانوادگی",
            "تاریخ",
            "ورود",
            "خروج",
            "وضعیت"
        ]

        # ==================================================
        # COLUMN RATIOS
        # ==================================================

        column_ratios = [
            1.15,
            1.35,
            1.15,
            1.05,
            1.05,
            1.10
        ]

        ratio_sum = sum(
            column_ratios
        )

        column_widths = [
            int(
                content_width
                * ratio
                / ratio_sum
            )
            for ratio in column_ratios
        ]

        column_widths[-1] += (
            content_width
            - sum(column_widths)
        )

        # ==================================================
        # HEIGHTS
        # ==================================================

        header_height = mm(12)
        row_height = mm(11)

        # ==================================================
        # FONTS
        # ==================================================

        header_font = painter.font()

        header_font.setPointSize(9)
        header_font.setBold(True)

        body_font = painter.font()

        body_font.setPointSize(8)
        body_font.setBold(False)

        # ==================================================
        # HEADER
        # ==================================================

        painter.setFont(
            header_font
        )

        painter.setPen(
            QColor("#1E3A5F")
        )

        painter.setBrush(
            QColor("#EAF3FF")
        )

        x = content_left

        for column, header in enumerate(
            headers
        ):

            rect = QRect(
                x,
                int(table_top),
                column_widths[column],
                header_height
            )

            painter.drawRect(
                rect
            )

            painter.drawText(
                rect,
                Qt.AlignmentFlag.AlignCenter
                | Qt.AlignmentFlag.AlignVCenter,
                header
            )

            x += column_widths[column]

        # ==================================================
        # BODY
        # ==================================================

        y = (
            table_top
            + header_height
        )

        row_index = 0

        while row_index < len(
            self.filtered_data
        ):

            # ==================================================
            # NEW PAGE
            # ==================================================

            if (
                y + row_height
                > content_bottom
                and is_printer
                and printer is not None
            ):

                printer.newPage()

                new_page = printer.pageRect(
                    QPrinter.Unit.DevicePixel
                )

                left = new_page.left()
                top = new_page.top()

                width = new_page.width()
                height = new_page.height()

                content_left = (
                    left + mm(12)
                )

                content_width = (
                    width
                    - mm(12)
                    - mm(12)
                )

                content_bottom = (
                    top
                    + height
                    - mm(12)
                )

                column_widths = [
                    int(
                        content_width
                        * ratio
                        / ratio_sum
                    )
                    for ratio in column_ratios
                ]

                column_widths[-1] += (
                    content_width
                    - sum(column_widths)
                )

                table_top = (
                    top
                    + mm(12)
                    + mm(8)
                )

                y = table_top

                # ------------------------------------------
                # HEADER ON NEW PAGE
                # ------------------------------------------

                painter.setFont(
                    header_font
                )

                painter.setPen(
                    QColor("#1E3A5F")
                )

                painter.setBrush(
                    QColor("#EAF3FF")
                )

                x = content_left

                for column, header in enumerate(
                    headers
                ):

                    rect = QRect(
                        x,
                        int(y),
                        column_widths[column],
                        header_height
                    )

                    painter.drawRect(
                        rect
                    )

                    painter.drawText(
                        rect,
                        Qt.AlignmentFlag.AlignCenter
                        | Qt.AlignmentFlag.AlignVCenter,
                        header
                    )

                    x += column_widths[column]

                y += header_height

            # ==================================================
            # CURRENT ROW
            # ==================================================

            row = self.filtered_data[
                row_index
            ]

            # ==================================================
            # ROW BACKGROUND
            # ==================================================

            painter.setBrush(
                QColor("#FFFFFF")
            )

            painter.setPen(
                QColor("#E2E8F0")
            )

            # ==================================================
            # CELLS
            # ==================================================

            x = content_left

            for column in range(
                len(headers)
            ):

                rect = QRect(
                    x,
                    int(y),
                    column_widths[column],
                    row_height
                )

                painter.drawRect(
                    rect
                )

                x += column_widths[column]

            # ==================================================
            # TEXT
            # ==================================================

            painter.setFont(
                body_font
            )

            painter.setPen(
                QColor("#334155")
            )

            x = content_left

            for column in range(
                len(headers)
            ):

                value = ""

                if column < len(row):

                    value = str(
                        row[column]
                    )

                rect = QRect(
                    x + mm(1),
                    int(y),
                    column_widths[column] - mm(2),
                    row_height
                )

                painter.drawText(
                    rect,
                    Qt.AlignmentFlag.AlignCenter
                    | Qt.AlignmentFlag.AlignVCenter,
                    value
                )

                x += column_widths[column]

            y += row_height

            row_index += 1

    # ==========================================================
    # GLOBAL EVENT FILTER
    # ==========================================================

    def eventFilter(
        self,
        watched,
        event
    ):

        if (
            event.type()
            == QEvent.Type.MouseButtonPress
        ):

            is_search_widget = (
                watched is self.toolbar.search_input
                or watched is self.toolbar.search_input.line_edit
                or (
                    isinstance(
                        watched,
                        QWidget
                    )
                    and self.toolbar.search_input.isAncestorOf(
                        watched
                    )
                )
            )

            if not is_search_widget:

                self.toolbar.search_input.clearFocus()

        return super().eventFilter(
            watched,
            event
        )

    # ==========================================================
    # MOUSE PRESS
    # ==========================================================

    def mousePressEvent(
        self,
        event
    ):

        self.toolbar.search_input.clearFocus()

        super().mousePressEvent(
            event
        )